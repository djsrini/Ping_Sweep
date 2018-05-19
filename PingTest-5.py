import os
import socket
import ipaddress
import subprocess
import time
from openpyxl import Workbook

name = str(time.time()).split('.')[0]

state = "b"
count = -1
input1 = []
cellId = 1

book = Workbook()
sheet = book.active
sheet['a1'] = "IP Address"
sheet['b1'] = "Ping Status"
sheet['c1'] = "DNS Name"

while state != "f":
    count = count + 1
    temp = input("Enter the IP:")
    if temp == "f":
        state = "f"
    else:
        input1.append(temp)

print(input1)

for i in input1:
    ipTemp = i
    print(ipTemp)
    ipSplitter = ipTemp.split(" ")
    print(ipSplitter)
    val1 = ipSplitter[0].split('.')
    val2 = ipSplitter[1].split('.')
    diff = int(val2[3]) - int(val1[3])
    print("The Difference is: " + str(diff))

    ipStr = str(ipSplitter[0])
    ip = ipaddress.IPv4Address(ipStr)

    for j in range(diff+1):
        ipi = ip + j
        cellId = cellId + 1
        ins = 'ping ' + str(ipi)
        ping = subprocess.Popen(["ping.exe", str(ipi)], stdout=subprocess.PIPE).communicate()[0]
        sheet.cell(row=cellId, column=1).value = str(ipi)
        if (('unreachable' in str(ping)) or ('Request timed out' in str(ping))):
            sheet.cell(row=cellId, column=2).value = "Not Pinging"
        else:
            sheet.cell(row=cellId, column=2).value = "Pinging"
        try:
            output = socket.gethostbyaddr(str(ipi))
            sheet.cell(row=cellId, column=3).value = str(output[0])
            print(output[0])
        except OSError:
            sheet.cell(row=cellId, column=3).value = "DNS Unavailable"
            print("DNS Unavailable")
        print(ping)

book.save(name + ".xlsx")


